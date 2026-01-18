import openpyxl
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "淘宝ESP32链接"

headers = ["序号", "商品标题", "链接"]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

products = [
    ["1", "ESP32-S3开发板乐鑫科技Espressif Systems 通信IC嵌入式", "https://www.taobao.com/list/item/VWpmSGNPRVR5TW9HeFdlU0dWc0tsQT09.htm"],
    ["2", "ESP32开发板乐鑫科技Core board ESP32-DevKitC开发板", "https://www.taobao.com/list/item/b3Z2djVxRit1N1BPUUVSZ0Rud2taUT09.htm"],
    ["3", "微雪ESP32-S3开发板支援ESP-IDF WiFi蓝牙学习板", "https://www.taobao.com/list/item/clYxSjBlandGSjVqL2NpTGdVOG4zZz09.htm"],
    ["4", "ESP32开发板带屏幕", "https://www.taobao.com/chanpin/9aa1fbc97157d6b9bfa4b722cb3da170f0269d922d6d61c8acc6a1c126ed585a.html"],
    ["5", "ESP32-S3开发套件乐鑫科技AIOT应用评估板学习开发板", "https://www.taobao.com/list/item/TTFCanVaMm9VZ214S0d3WVZ1a0VkUT09.htm"],
    ["6", "ESP32物联网开发套件 高配进阶版蓝牙WiFi模组 适用Arduino学习板", "https://www.taobao.com/list/item/a2lUZlJwWGRjZXVvNlVBVmI4Z0R0UT09.htm"],
    ["7", "ESP-32S WiFi开发板 支援蓝牙与串口WiFi功能 NodeMCU ESP32模组", "https://www.taobao.com/list/item/674698379065.htm"],
    ["8", "ESP32-S3核心板 乐鑫ESP32-S3 开发板学习板 采用N8R8", "https://www.taobao.com/list/item/QnlNb3FrZ3lhc0NuTzJnV2d6TmVidz09.htm"],
    ["9", "esp32开发板促销价格", "https://www.taobao.com/chanpin/1231994a07dc10ab579a3e80158ac063.html"],
    ["10", "源地YD-ESP32开发板WROOM-32E核心板乐鑫WIFI蓝牙", "https://www.taobao.com/list/item/675317800762.htm"]
]

for row_idx, product in enumerate(products, 2):
    for col_idx, value in enumerate(product, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 80
ws.column_dimensions['C'].width = 80

wb.save('taobao_esp32_links.xlsx')
print("Excel文件已创建: taobao_esp32_links.xlsx")